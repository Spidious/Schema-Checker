import openpyxl

# Parser objet for handling xlsx files
class xlsxParser(object):
    name = ""
    email = ""
    data = {}
    def __init__(self, file: str):
        # If the input file is not empty, parse the file
        # if file!=None:
        f = file.split('_')
        self.name = f[0]
        self.email = f[1].replace(".xlsx", "")

        if file!=None:
            #  Define variable to load the dataframe
            dataframe = openpyxl.load_workbook(file)
            # Define variable to read sheet
            df = dataframe.active
            

            output = {}
            # for each column, use the element as a key, the rest as the value
            for col in df.iter_cols(1, df.max_column):
                self.data[col[0].value] = [cell.value for cell in col[1:]]

    def get_headers(self) -> list:
        # Loop through each key in self.data appending it to headers
        headers = []
        for key in self.data.keys():
            headers.append(key)
        # return headers
        return headers


if __name__ == "__main__":
    reader = xlsxParser("userdb_example@gmail.com.xlsx")
    pass